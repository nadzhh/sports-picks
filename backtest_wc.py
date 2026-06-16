"""
backtest_wc.py — Backtesting des picks WC pour mesurer l'impact des règles.

Lit data/picks_history.json filtré WC, calcule WR / ROI par bucket conf et
par dimension (avant/après chaque règle introduite), génère un Excel
exploitable.

Sortie : data/backtest_wc.xlsx avec onglets :
  - Picks      : 1 ligne par pick WC avec ses méta + result + edge + tier
  - Bucket     : WR/ROI par bucket de confidence (0-40, 40-50, ..., 90-100)
  - Family     : WR/ROI par famille (1X2, DC, TOTAL, BTTS, BUT_TEAM, SCORE)
  - Tier       : WR/ROI par tier (safe / ok / fun)
  - Team       : WR/ROI par équipe (home/away combinés)
  - Summary    : ROI/WR global + recommandations

Usage : python backtest_wc.py
"""
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[X] openpyxl manquant. Installe : pip install openpyxl")
    raise SystemExit(1)

HISTORY_FILE = Path("data/picks_history.json")
OUT_FILE     = Path("data/backtest_wc.xlsx")


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _est_cote(p):
    """Estimation de cote depuis la proba (margin book 6%)."""
    if not p or p <= 0: return None
    return round((1.0 / p) * 0.94, 2)


def _effective_cote(pick):
    """Cote book si dispo, sinon cote_min, sinon estimation depuis confidence."""
    c = pick.get("cote")
    if c and c > 0: return float(c)
    cm = pick.get("cote_min")
    if cm and cm > 0: return float(cm)
    conf = pick.get("confidence", 0)
    return _est_cote(conf / 100.0) if conf else None


def _is_wc_pick(p):
    """True si le pick est sur un match WC."""
    league = (p.get("league") or "").lower()
    if "world cup" in league or "coupe du monde" in league: return True
    dir_str = (p.get("direction") or "").lower()
    if dir_str.startswith("wc_"): return True
    return False


def _family(direction):
    """Détermine la famille de marché du pick."""
    d = (direction or "").lower()
    if d.startswith("wc_score_"): return "SCORE_EXACT"
    if "1x2" in d or "home_win" in d or "away_win" in d or "draw" in d: return "1X2"
    if "dc_" in d or "double" in d: return "DC"
    if "over" in d or "under" in d: return "TOTAL"
    if "btts" in d: return "BTTS"
    if "scores_" in d or "no_score" in d or "team_score" in d: return "BUT_TEAM"
    if "clean_sheet" in d: return "CLEAN_SHEET"
    return "OTHER"


def _bucket(conf):
    """Bucket de confidence [0-40, 40-50, 50-60, 60-70, 70-80, 80-90, 90-100]."""
    if conf is None: return "?"
    if conf < 40: return "<40"
    if conf < 50: return "40-49"
    if conf < 60: return "50-59"
    if conf < 70: return "60-69"
    if conf < 80: return "70-79"
    if conf < 90: return "80-89"
    return "90+"


def _payout(result, cote):
    """Renvoie le gain net pour 1 unité misée (PnL)."""
    if cote is None: return 0
    if result == "WIN":  return cote - 1
    if result == "LOSS": return -1
    if result == "PUSH": return 0
    if result == "DNP":  return 0
    return 0  # PENDING ignore


# ─── Compute stats ───────────────────────────────────────────────────────────

def compute_aggregates(picks):
    """Compute WR / ROI agrégés par dimension."""
    by_bucket = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0, "push": 0, "pnl": 0.0})
    by_family = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0, "push": 0, "pnl": 0.0})
    by_tier   = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0, "push": 0, "pnl": 0.0})
    by_team   = defaultdict(lambda: {"n": 0, "wins": 0, "losses": 0, "push": 0, "pnl": 0.0})
    total = {"n": 0, "wins": 0, "losses": 0, "push": 0, "pnl": 0.0}

    for p in picks:
        result = (p.get("result") or "").upper()
        if result not in ("WIN", "LOSS", "PUSH", "DNP"): continue
        cote = _effective_cote(p)
        conf = p.get("confidence", 0)
        pnl  = _payout(result, cote)
        bucket = _bucket(conf)
        family = _family(p.get("direction"))
        tier   = p.get("tier") or "?"
        # Équipes : home + away depuis matchup "X vs Y"
        matchup = p.get("matchup") or ""
        teams = [t.strip() for t in matchup.split(" vs ")] if " vs " in matchup else [matchup]
        for bag, key in [(by_bucket, bucket), (by_family, family), (by_tier, tier)]:
            bag[key]["n"] += 1
            if result == "WIN":  bag[key]["wins"] += 1
            if result == "LOSS": bag[key]["losses"] += 1
            if result == "PUSH": bag[key]["push"] += 1
            bag[key]["pnl"] += pnl
        for t in teams:
            by_team[t]["n"] += 1
            if result == "WIN":  by_team[t]["wins"] += 1
            if result == "LOSS": by_team[t]["losses"] += 1
            if result == "PUSH": by_team[t]["push"] += 1
            by_team[t]["pnl"] += pnl
        total["n"] += 1
        if result == "WIN":  total["wins"] += 1
        if result == "LOSS": total["losses"] += 1
        if result == "PUSH": total["push"] += 1
        total["pnl"] += pnl

    return total, by_bucket, by_family, by_tier, by_team


# ─── Write Excel ─────────────────────────────────────────────────────────────

HEAD_FONT = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
WIN_FILL  = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
LOSS_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
PUSH_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
PEND_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")


def _write_header(ws, row, cols, widths=None):
    for i, label in enumerate(cols):
        c = ws.cell(row=row, column=i + 1, value=label)
        c.font = HEAD_FONT
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w


def _wr_str(wins, losses, push):
    real = wins + losses
    if real == 0: return ""
    return f"{round(wins / real * 100, 1)}%"


def _roi_str(pnl, n):
    if n == 0: return ""
    return f"{round(pnl / n * 100, 1)}%"


def write_workbook(picks, total, by_bucket, by_family, by_tier, by_team):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.cell(1, 1, value=f"Backtest Coupe du Monde 2026 — généré {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(bold=True, size=14)
    ws.cell(3, 1, value="Picks WC résolus").font = Font(bold=True)
    ws.cell(3, 2, value=total["n"])
    ws.cell(4, 1, value="Wins").font = Font(bold=True)
    ws.cell(4, 2, value=total["wins"])
    ws.cell(5, 1, value="Losses").font = Font(bold=True)
    ws.cell(5, 2, value=total["losses"])
    ws.cell(6, 1, value="Pushes").font = Font(bold=True)
    ws.cell(6, 2, value=total["push"])
    real = total["wins"] + total["losses"]
    wr = round(total["wins"] / real * 100, 1) if real else 0
    ws.cell(7, 1, value="Win Rate").font = Font(bold=True)
    ws.cell(7, 2, value=f"{wr}%")
    roi = round(total["pnl"] / total["n"] * 100, 2) if total["n"] else 0
    ws.cell(8, 1, value="ROI (%)").font = Font(bold=True)
    ws.cell(8, 2, value=f"{roi}%")
    ws.cell(8, 2).fill = WIN_FILL if roi > 0 else (LOSS_FILL if roi < 0 else PUSH_FILL)
    ws.cell(10, 1, value="PnL net (unités)").font = Font(bold=True)
    ws.cell(10, 2, value=round(total["pnl"], 2))

    # Top recommandations
    ws.cell(13, 1, value="Top 3 buckets les plus rentables").font = Font(bold=True)
    bucket_sorted = sorted(by_bucket.items(), key=lambda x: -(x[1]["pnl"]/x[1]["n"] if x[1]["n"] else -999))[:3]
    for i, (b, s) in enumerate(bucket_sorted):
        roi_b = round(s["pnl"]/s["n"]*100,1) if s["n"] else 0
        ws.cell(14+i, 1, value=f"  conf {b}").font = Font(italic=True)
        ws.cell(14+i, 2, value=f"ROI {roi_b}% ({s['n']} picks)")

    ws.cell(18, 1, value="Top 3 familles les plus rentables").font = Font(bold=True)
    fam_sorted = sorted(by_family.items(), key=lambda x: -(x[1]["pnl"]/x[1]["n"] if x[1]["n"] else -999))[:3]
    for i, (f, s) in enumerate(fam_sorted):
        roi_f = round(s["pnl"]/s["n"]*100,1) if s["n"] else 0
        ws.cell(19+i, 1, value=f"  {f}").font = Font(italic=True)
        ws.cell(19+i, 2, value=f"ROI {roi_f}% ({s['n']} picks)")

    ws.cell(23, 1, value="Familles à VIRER (ROI < -20% sur N≥5)").font = Font(bold=True, color="DC2626")
    bad = [(f, s) for f, s in by_family.items() if s["n"] >= 5 and s["pnl"]/s["n"] < -0.20]
    for i, (f, s) in enumerate(bad):
        ws.cell(24+i, 1, value=f"  {f}").font = Font(italic=True)
        ws.cell(24+i, 2, value=f"ROI {round(s['pnl']/s['n']*100,1)}% ({s['n']} picks)")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 32

    # ── Picks ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Picks")
    cols = ["Date", "Match", "Direction", "Label", "Family", "Tier", "Conf%",
            "Cote (eff)", "Cote source", "Edge_pp", "Result", "PnL"]
    widths = [12, 36, 18, 38, 14, 8, 7, 11, 12, 9, 8, 9]
    _write_header(ws, 1, cols, widths)
    row = 2
    for p in sorted(picks, key=lambda x: (x.get("date") or "", x.get("matchup") or "")):
        result = (p.get("result") or "").upper()
        cote = _effective_cote(p)
        src = "book" if p.get("cote") else ("cote_min" if p.get("cote_min") else "simulé")
        edge = p.get("edge_pp")
        ws.cell(row, 1, value=p.get("date") or "")
        ws.cell(row, 2, value=p.get("matchup") or "")
        ws.cell(row, 3, value=p.get("direction") or "")
        ws.cell(row, 4, value=p.get("label") or "")
        ws.cell(row, 5, value=_family(p.get("direction")))
        ws.cell(row, 6, value=p.get("tier") or "")
        ws.cell(row, 7, value=p.get("confidence") or "")
        ws.cell(row, 8, value=cote if cote else "")
        ws.cell(row, 9, value=src)
        ws.cell(row, 10, value=edge if edge is not None else "")
        ws.cell(row, 11, value=result)
        ws.cell(row, 12, value=round(_payout(result, cote), 2))
        # Couleur ligne
        fill = {"WIN": WIN_FILL, "LOSS": LOSS_FILL, "PUSH": PUSH_FILL,
                "PENDING": PEND_FILL, "DNP": PUSH_FILL}.get(result, None)
        if fill:
            for col in range(1, 13):
                ws.cell(row, col).fill = fill
        row += 1

    # ── Bucket / Family / Tier / Team sheets ────────────────────────────────
    def _write_bag(ws_name, label_col, bag, sort_by="pnl"):
        ws = wb.create_sheet(ws_name)
        cols = [label_col, "N picks", "Wins", "Losses", "Push", "Win Rate", "ROI %", "PnL"]
        _write_header(ws, 1, cols, [22, 9, 7, 8, 7, 11, 9, 9])
        # Tri par PnL/N décroissant
        items = sorted(bag.items(), key=lambda x: -(x[1]["pnl"]/x[1]["n"] if x[1]["n"] else -999))
        for i, (k, s) in enumerate(items, start=2):
            ws.cell(i, 1, value=k)
            ws.cell(i, 2, value=s["n"])
            ws.cell(i, 3, value=s["wins"])
            ws.cell(i, 4, value=s["losses"])
            ws.cell(i, 5, value=s["push"])
            ws.cell(i, 6, value=_wr_str(s["wins"], s["losses"], s["push"]))
            roi_pct = s["pnl"]/s["n"]*100 if s["n"] else 0
            ws.cell(i, 7, value=round(roi_pct, 1))
            ws.cell(i, 8, value=round(s["pnl"], 2))
            if roi_pct > 5:
                for c in range(1, 9): ws.cell(i, c).fill = WIN_FILL
            elif roi_pct < -10:
                for c in range(1, 9): ws.cell(i, c).fill = LOSS_FILL

    _write_bag("Bucket conf", "Bucket", by_bucket)
    _write_bag("Family", "Famille", by_family)
    _write_bag("Tier", "Tier", by_tier)
    _write_bag("Team", "Équipe", by_team)

    wb.save(OUT_FILE)
    return OUT_FILE


# ─── Main ────────────────────────────────────────────────────────────────────

def run():
    if not HISTORY_FILE.exists():
        print(f"[X] {HISTORY_FILE} introuvable")
        return
    history = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
    picks = history.get("picks", [])
    wc_picks = [p for p in picks if _is_wc_pick(p)]
    resolved = [p for p in wc_picks if (p.get("result") or "").upper() in ("WIN", "LOSS", "PUSH", "DNP")]
    pending  = [p for p in wc_picks if (p.get("result") or "").upper() in ("", "PENDING")]
    print(f"=== Backtest WC : {len(wc_picks)} picks WC ({len(resolved)} résolus, {len(pending)} pending) ===")

    if not resolved:
        print("[!] Aucun pick WC résolu — rien à backtester")
        # Génère quand même un Excel vide pour vérifier le format
    total, by_bucket, by_family, by_tier, by_team = compute_aggregates(resolved)

    out = write_workbook(resolved, total, by_bucket, by_family, by_tier, by_team)
    print(f"[OK] Excel généré : {out}")
    real = total["wins"] + total["losses"]
    wr = round(total["wins"] / real * 100, 1) if real else 0
    roi = round(total["pnl"] / total["n"] * 100, 2) if total["n"] else 0
    print(f"  Global : WR {wr}% · ROI {roi}% · PnL net {round(total['pnl'],2)}u")


if __name__ == "__main__":
    run()
